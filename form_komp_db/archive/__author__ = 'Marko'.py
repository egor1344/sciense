__author__ = 'Marko'

import sys
import sqlite3

import openpyxl
from PyQt5.QtWidgets import (QMainWindow, QAction,
                             QFileDialog, QApplication, QLabel,
                             QComboBox, QGridLayout, QTableView, QWidget)
from PyQt5.QtGui import QStandardItemModel
from PyQt5.QtCore import QModelIndex


class Example(QMainWindow):
    def __init__(self):
        super().__init__()

        self.initUI()

    def initUI(self):
        self.resize(800, 600)

        centralWidget = QWidget(self)
        grid = QGridLayout(centralWidget)
        grid.setSpacing(10)

        self.listOfDiscipline = QComboBox(centralWidget)
        grid.addWidget(self.listOfDiscipline, 0, 0, 1, 1)

        textIn = QLabel(centralWidget)
        textIn.setText('Входные компетенции')
        grid.addWidget(textIn, 1, 0, 1, 1)

        self.tableIn = QTableView(centralWidget)
        grid.addWidget(self.tableIn, 2, 0, 1, 1)

        textOut = QLabel(centralWidget)
        textOut.setText('Выходные компетенции')
        grid.addWidget(textOut, 3, 0, 1, 1)

        self.tableOut = QTableView(centralWidget)
        grid.addWidget(self.tableOut, 4, 0, 1, 1)

        self.setCentralWidget(centralWidget)

        openFile = QAction('Открыть базу данных', self)
        openFile.setShortcut('Ctrl+O')
        openFile.setStatusTip('Открыть базу данных комптенций')
        openFile.triggered.connect(self.MainFunc)

        exitWindow = QAction('Выход', self)
        exitWindow.setStatusTip('Завершение программы')
        exitWindow.triggered.connect(self.close)

        saveFile_one = QAction('Сохранить дисциплину', self)
        saveFile_one.setShortcut('Ctrl+S')
        saveFile_one.setStatusTip('Сохранить одну дисциплину в формате xlsx')
        saveFile_one.triggered.connect(self.save_one)

        saveFile_all = QAction('Сохранить все дисциплины', self)
        saveFile_all.setShortcut('Ctrl+A')
        saveFile_all.setStatusTip('Сохранить все дисциплины в формате xlsx')
        saveFile_all.triggered.connect(self.save_all)

        menubar = self.menuBar()
        fileMenu = menubar.addMenu('&Файл')
        fileMenu.addAction(openFile)
        fileSaveMenu = fileMenu.addMenu('&Дисциплины')
        fileSaveMenu.addAction(saveFile_one)
        fileSaveMenu.addAction(saveFile_all)
        fileMenu.addAction(fileSaveMenu.menuAction())
        fileMenu.addSeparator()
        fileMenu.addAction(exitWindow)

        self.nameFileOpen = ''

        self.statusBar()
        self.setWindowTitle('Формирование компетенций')
        self.show()

    def MainFunc(self):
        """
        Основная функция
        :return:
        """
        self.nameFileOpen = self.showDialog()
        self.listOfDiscipline.activated[str].connect(self.comboPrint)

    def comboPrint(self, displ):
        self.formInComp(displ, self.nameFileOpen, 'table')

    def showDialog(self):
        """
        Открытие базы данных
        :return:
        """
        fname = QFileDialog.getOpenFileName(self, 'Открыть базу данных', '/', 'DataBase(*.sqlite)')[0]
        self.displ = self.listDB(fname)
        return fname

    def listDB(self, namebd):
        """
        Получение списка таблиц базы данных
        :param nameDB:
        :return:
        """
        cache = []
        db = sqlite3.connect(namebd)
        cur = db.cursor()

        k = cur.execute('SELECT * FROM sqlite_master WHERE type = "table"')
        for j in k:
            if j[1] != 'sqlite_stat1' and j[1] != 'ОК' and j[1] != 'ОПК':
                cache.append(j[1])

        cur.close()
        db.close()
        self.listOfDiscipline.addItems(cache)
        return cache

    def formInComp(self, displ, namebd, type):
        """
        Формирование входящих компетенций на основе выходящих
        :param list, namebd:
        :return:
        """
        self.cacheIn = []
        self.cacheOut = []
        self.cacheInDisp = []
        rowIn = 0
        rowOut = 0
        db = sqlite3.connect(namebd)
        cur = db.cursor()
        k = cur.execute('SELECT * FROM ' + displ)
        for line in k:
            inDisp = self.procInDisp(line[6])
            if str(line[0]).find('К') != -1:
                rowOut = rowOut + 1
                self.cacheOut.append(line)
            if inDisp != '':
                self.cacheInDisp.append(inDisp)

        for dis in self.cacheInDisp:
            if dis in self.displ:
                k = cur.execute('SELECT * FROM ' + dis)
                for line in k:
                    if (line[0] != '') and (line[0] != '\t'):
                        self.cacheIn.append(line)
                        rowIn = rowIn + 1

        self.cacheIn.sort()
        self.cacheOut.sort()
        if type == 'table':
            self.createTable('out', rowOut, self.cacheOut)
            self.createTable('in', rowIn, self.cacheIn)
        else:
            return self.cacheIn, self.cacheOut

        cur.close()
        db.close()

    def procInDisp(self, disp):
        """
        Обработка входящих дисциплин для применения в SQL запросах
        :param disp:
        :return:
        """
        m = disp
        m = m.strip()
        m = m.replace(' ', '_')
        m = m.replace('  ', '')
        m = m.replace('   ', '')
        m = m.replace('-', '_')
        m = m.replace('.', '')
        m = m.replace('«', '')
        m = m.replace('»', '')
        m = m.replace(',', '')
        m = m.capitalize()
        return m

    def createTable(self, type_table, rows, list_of_komp):
        """
        Создание и заполнение таблицы
        :param type_table:
        :param row:
        :param list_komp:
        :param list_desc:
        :return:
        """
        if type_table == 'out':
            model = QStandardItemModel(rows, 2)
            list_horizontal = ['Код компетенции', 'Описание компетенции']
            model.setHorizontalHeaderLabels(list_horizontal)
            self.tableOut.setModel(model)
            self.tableOut.setColumnWidth(0, 110)
            self.tableOut.setColumnWidth(1, self.tableOut.width() - 130)
            for row in range(rows):
                for column in range(2):
                    index = model.index(row, column, QModelIndex())
                    model.setData(index, str(list_of_komp[row][column]).strip())
        else:
            model = QStandardItemModel(rows, 2)
            list_horizontal = ['Код компетенции', 'Описание компетенции']
            model.setHorizontalHeaderLabels(list_horizontal)
            self.tableIn.setModel(model)
            self.tableIn.setColumnWidth(0, 110)
            self.tableIn.setColumnWidth(1, self.tableIn.width() - 130)
            for row in range(rows):
                for column in range(2):
                    index = model.index(row, column, QModelIndex())
                    model.setData(index, str(list_of_komp[row][column]).strip())

    def save_one(self):
        wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = str(self.listOfDiscipline.currentText())[:20]
        ws.append(['Входящие компетенции'])
        for line in self.cacheIn:
            ws.append([line[0].strip(), line[1].strip()])
        ws.append(['Выходящие компетенции'])
        for line in self.cacheOut:
            ws.append([line[0].strip(), line[1].strip()])

        wb.save(self.listOfDiscipline.currentText() + '.xlsx')

    def save_all(self):
        count_displ = self.listOfDiscipline.count()
        wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = 'Дисциплины'
        ws['A1'] = 'Сохранениые дисциплины'
        for count in range(count_displ):
            ws.append([str(self.listOfDiscipline.itemText(count))])
        for count in range(count_displ):
            inDis, outDis = self.formInComp(self.listOfDiscipline.itemText(count), self.nameFileOpen, ' ')
            ws = wb.create_sheet(self.listOfDiscipline.itemText(count)[:20])
            ws.append(['Входящие компетенции'])
            for line in inDis:
                ws.append([line[0].strip(), line[1].strip()])
            ws.append(['Выходящие компетенции'])
            for line in outDis:
                ws.append([line[0].strip(), line[1].strip()])
        wb.save('Все дисциплины.xlsx')


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = Example()
    sys.exit(app.exec_())